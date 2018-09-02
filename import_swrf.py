import argparse
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
import openpyxl.utils as xlutils
import sqlite3

def init_db(filename):
  conn = sqlite3.connect(filename)

  with open('swrf_schema.sql') as f:
    conn.cursor().executescript(f.read())

  conn.commit()

  return conn

def find_boat_rows(ws, curr_row):
  if not ws.cell(row=curr_row, column=1).value:
    print('ERROR: missing boat name where expected')
    return

  curr_row += 1
  row_cnt = 1

  # wait until we find an undefined field
  while ws.cell(row=curr_row, column=1).value:
    curr_row += 1
    row_cnt += 1

  while not ws.cell(row=curr_row, column=1).value:
    curr_row += 1
    row_cnt += 1

  return row_cnt

def skipper_lookup(person, db_curs):
 
  if not person:
    first_name = 'UNKNOWN'
    last_name = None
  else:
    person_split = person.split(' ')
    if len(person_split) == 2:
      first_name = person_split[0]
      last_name  = person_split[1]
    else:
      first_name = person
      last_name  = None

  result = db_curs.execute('select personid from person where first_name=? and last_name=? limit 1', (first_name, last_name)).fetchone()

  if(not result):
    db_curs.execute('insert into person(first_name, last_name) values (?, ?)', (first_name, last_name))

    if last_name:
      result = db_curs.execute('select personid from person where first_name=? and last_name=? limit 1', (first_name, last_name)).fetchone()
    else:
      result = db_curs.execute('select personid from person where first_name=? and last_name is null limit 1', (first_name,)).fetchone()

  return result[0]

def parse_races(ws, curr_row, db_curs, boatid, last_race_col):
  
  for raceid, race in enumerate(reversed(list(data.iter_cols(min_row=curr_row, max_row=curr_row+2, max_col=last_race_col, min_col=xlutils.column_index_from_string('I'))))):
    start_time = race[0].value
    finish_time = race[1].value

    rc_list = ['RC', 'R/C']

    if start_time in rc_list or finish_time in rc_list:
      result = 'RC'
      start_time = None
      finish_time = None
    elif start_time == 'OCS' or finish_time == 'OCS':
      result = 'OCS'
      start_time = None
      finish_time = None
    elif start_time == 'DNF' or finish_time == 'DNF':
      result = 'DNF'
      start_time = None
      finish_time = None
    elif not start_time:
      result = None
    else:
      result = 'FINISH'
      start_time = str(start_time)
      finish_time = str(finish_time)

    if result:
      db_curs.execute('insert into race_result(raceid, boatid, result, start_time, finish_time) values (?, ?, ?, ?, ?)', (raceid+1, boatid, result, start_time, finish_time))
    
    
    
def decode_format_a(ws, curr_row, db_curs, last_race_col):
  cell = ws.cell(row=curr_row, column=1)
  name = cell.value
  skipper = cell.offset(0, 1).value
  sail_num = cell.offset(0, 2).value
  non_spin = cell.offset(0, 3).value
  model = cell.offset(1, 5).value
  phrf_rlc = cell.offset(2, 5).value
  phrf_buoy = cell.offset(2, 6).value
  meas_i = cell.offset(3, 5).value
  meas_j = cell.offset(3, 6).value
  meas_p = cell.offset(4, 5).value
  meas_e = cell.offset(4, 6).value

  if non_spin == "NS" or non_spin == "Y":
    non_spin = 1
  else:
    non_spin = 0

  skipper_id = skipper_lookup(skipper, db_curs)

  db_curs.execute('insert into boat(skipperid, name, sail_number, model, meas_i, meas_j, meas_p, meas_e, phrf_rlc, phrf_buoy, nonspin) values (?, ?, ?, ?, ?, ?, ?, ?, ?, ? , ?)', (skipper_id, name, sail_num, model, meas_i, meas_j, meas_p, meas_e, phrf_rlc, phrf_buoy, non_spin))

  boat_id = db_curs.execute('select boatid from boat where rowid=?', (db_curs.lastrowid,)).fetchone()[0]

  parse_races(ws, curr_row, db_curs, boat_id, last_race_col)

def decode_format_b(ws, curr_row, db_curs, last_race_col):
  decode_format_a(ws, curr_row, db_curs, last_race_col)

def decode_format_c(ws, curr_row, db_curs, last_race_col):
  cell = ws.cell(row=curr_row, column=1)
  name = cell.value
  skipper = cell.offset(1, 0).value
  sail_num = cell.offset(2, 0).value
  non_spin = cell.offset(3, 0).value

  # not all entries are consistent with ordering
  if type(skipper) == int:
    skipper, sail_num = sail_num, skipper

  if non_spin == "NS" or non_spin == "Y":
    non_spin = 1
  else:
    non_spin = 0

  skipper_id = skipper_lookup(skipper, db_curs)

  db_curs.execute('insert into boat(skipperid, name, sail_number, nonspin) values (?, ?, ?, ?)', (skipper_id, name, sail_num, non_spin))

  boat_id = db_curs.execute('select boatid from boat where rowid=?', (db_curs.lastrowid,)).fetchone()[0]

  parse_races(ws, curr_row, db_curs, boat_id, last_race_col)


parser = argparse.ArgumentParser(description='Parse SWRF spreadsheet to SQLite database.')
parser.add_argument('xlsx', type=str,
                    help='xlsx file')

args = parser.parse_args()

db_conn = init_db('swrf.db')
db_curs = db_conn.cursor()

wb = load_workbook(args.xlsx)

data = wb['Hcap']

oldest_race_column = None;

for race in reversed(list(data.iter_cols(min_row=1, max_row=8, min_col=xlutils.column_index_from_string('I')))):
  if race[0].value or race[4].value:

    if not oldest_race_column:
      oldest_race_column = race[0].col_idx

    if race[0].value:
      race_name = race[0].value
    else:
      race_name = 'UNKNOWN'

    if race[1].value:
      race_name += ' ' + str(race[1].value)

    if race[3].value:
      distance = race[3].value
    else:
      distance = None

    if race[4].value:
      date = race[4].value + timedelta(hours=12)
    else:
      date = 0

    db_curs.execute('insert into race(name, date_time, distance_nm) values (?, ?, ?)', (race_name, date, distance))

curr_row = 10

while(True):
  boat_rows = find_boat_rows(data, curr_row) 

  if boat_rows >= 19 and boat_rows <= 20:
    decode_format_a(data, curr_row, db_curs, oldest_race_column)
  elif boat_rows >= 6:
    if data.cell(row=curr_row, column=2).value:
      decode_format_b(data, curr_row, db_curs, oldest_race_column)
    else:
      decode_format_c(data, curr_row, db_curs, oldest_race_column)
  else:
    print('Unknown boat format found on row {}, rows = {}'.format(curr_row, boat_rows))

  curr_row += boat_rows

  if data.cell(row=curr_row, column=1).value.startswith('LAST'):
    break

db_conn.commit()





